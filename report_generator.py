import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json

DATA_DIR = "reports"
ANALYTICS_DIR = "analytics"
CONFIG_FILE = "bar_config.json"

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(ANALYTICS_DIR, exist_ok=True)

def load_config():
    try:
        with open(CONFIG_FILE, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {"bar_name": "LIQUOR WORLD", "tax_rate": 0.16, "currency": "Ksh"}

def get_full_inventory():
    """Returns the complete inventory with categories and subcategories"""
    inventory = {
        "BEERS": {
            "BOTTLES": [
                {"name": "HENNEKEN BOTTLE 330ML", "price": 300.00},
                {"name": "SAVANA DRY 330ML", "price": 300.00},
                {"name": "DESPARADO TEQ BEER 330ML", "price": 300.00},
                {"name": "SNAPP DRY", "price": 250.00},
                {"name": "HUNTERS 330ML", "price": 300.00},
                {"name": "MANYATTA", "price": 300.00},
                {"name": "BALOZI", "price": 250.00},
                {"name": "GUINESS", "price": 250.00},
                {"name": "TUSKER LITE", "price": 250.00},
                {"name": "WHITE CAP", "price": 250.00},
                {"name": "TUSKER LITE", "price": 250.00},
                {"name": "TUSKER LGR", "price": 250.00},
                {"name": "BLACK ICE", "price": 250.00},
                {"name": "TUSKER CIDER", "price": 250.00},
                {"name": "PILSNER", "price": 250.00},
                {"name": "WHITE LIGHT", "price": 350.00},
                {"name": "OJ 12%", "price": 0.00}
            ],
            "CANS": [
                {"name": "GORDONS CAN", "price": 270.00},
                {"name": "GUINESS CANS", "price": 270.00},
                {"name": "GUARANA 300ML", "price": 250.00},
                {"name": "FAXE", "price": 300.00},
                {"name": "TUSKER LITE CANS", "price": 270.00},
                {"name": "TUSKER CANS 500ML", "price": 270.00},
                {"name": "TUSKER CIDER CANS 500ML", "price": 270.00},
                {"name": "BLACK ICE", "price": 250.00},
                {"name": "BLACK ICE PUNCH", "price": 250.00},
                {"name": "TUSKER MALT CANS", "price": 270.00},
                {"name": "PILSNER CANS", "price": 250.00},
                {"name": "WHITE CAP LARGER CANS", "price": 270.00},
                {"name": "BALOZI CANS", "price": 270.00},
                {"name": "SNAPP APPLE CANS 300MLS", "price": 250.00},
                {"name": "GUINESS SMOOTH CANS", "price": 250.00},
                {"name": "WHITE CAP LITE CANS", "price": 270.00},
                {"name": "HEINEKEN CAN", "price": 350.00}
            ]
        },
        "ENERGY DRINK": {
            "": [
                {"name": "RED BULL 250ML", "price": 250.00},
                {"name": "MONSTER", "price": 250.00},
                {"name": "Predator", "price": 70.00}
            ]
        },
        "SOFT DRINKS": {
            "": [
                {"name": "DELMONTE 1LTR", "price": 300.00},
                {"name": "PET SODA 2LTR", "price": 200.00},
                {"name": "PET SODA 1.25ML", "price": 150.00},
                {"name": "LEMONADE", "price": 70.00},
                {"name": "PET SODA 500ML", "price": 100.00},
                {"name": "soda 330ml", "price": 70.00}
            ]
        },
        "WINES": {
            "": [
                {"name": "RICHOT 350ML", "price": 750.00},
                {"name": "RICHOT 750ML", "price": 1500.00},
                {"name": "VICEROY 750ML", "price": 1600.00},
                {"name": "VICEROY 250ML", "price": 650.00},
                {"name": "VICEROY 375ML", "price": 750.00}
            ]
        },
        "SPIRIT": {
            "": [
                {"name": "KONYANGI 750ML", "price": 1100.00},
                {"name": "KONYANGI 500ML", "price": 700.00},
                {"name": "KONYANGI 250ML", "price": 380.00},
                {"name": "COUNTY 750ML", "price": 1200.00},
                {"name": "COUNTY 250ML", "price": 350.00}
            ]
        },
        "TEQUILLA": {
            "": [
                {"name": "CAMINO 750ML", "price": 2700.00}
            ]
        },
        "COGNAC": {
            "": [
                {"name": "Martel vs 700ML", "price": 5600.00},
                {"name": "MARTEL VSOP 700ml", "price": 0.00},
                {"name": "HANNESY750ML cognac", "price": 5600.00}
            ]
        },
        "RUM": {
            "": [
                {"name": "CAPTAIN gold 750ml", "price": 1300.00},
                {"name": "CAPTAIN MORGAN 750ml SPICED", "price": 2200.00},
                {"name": "CAPTAIN MORGAN 1ltr SPICED", "price": 2200.00},
                {"name": "CAPTAIN MORGAN 1/4", "price": 500.00},
                {"name": "KENYA CANE 250ML", "price": 350.00},
                {"name": "KENYA CANE 350ML", "price": 450.00},
                {"name": "KENYA CANE 750ML", "price": 900.00}
            ]
        },
        "LIQUER": {
            "": [
                {"name": "orijin 750ml", "price": 850.00},
                {"name": "Orijin quarter", "price": 450.00},
                {"name": "V&A 250ML", "price": 400.00},
                {"name": "V&A 750ML", "price": 1100.00},
                {"name": "Southern Comfort 1 ltr", "price": 2500.00},
                {"name": "Southern Comfort 750ml", "price": 2200.00},
                {"name": "Southern Comfort 750ml black", "price": 1900.00},
                {"name": "JEGERMISTER 1l", "price": 0.00},
                {"name": "JEGERMISTER 750ML", "price": 2500.00},
                {"name": "BEST cream 750ML", "price": 1200.00},
                {"name": "Amarula 1ltr", "price": 3000.00},
                {"name": "Amarula 750ml", "price": 2500.00},
                {"name": "Amarula 1/2", "price": 1700.00},
                {"name": "BAILEYS 1LTR", "price": 3400.00},
                {"name": "BAILEYS 750ml", "price": 2800.00},
                {"name": "BAILEYS 1/2", "price": 2000.00},
                {"name": "BAILEYS DELIGHT 750ML", "price": 1200.00}
            ]
        },
        "ROOSTER": {
            "": [
                {"name": "PALL MALL", "price": 20.00},
                {"name": "EMBASY", "price": 0.00},
                {"name": "DUNHIL", "price": 30.00},
                {"name": "SPORTSMAN", "price": 30.00},
                {"name": "vello small", "price": 25.00},
                {"name": "vello medim", "price": 300.00},
                {"name": "", "price": 350.00}
            ]
        },
        "TOTS": {
            "": [
                {"name": "JACKDANIEL TOT", "price": 200.00},
                {"name": "GILBEYS TOT", "price": 150.00},
                {"name": "JERGERMEISTER TOT", "price": 200.00},
                {"name": "CAMINO TOT", "price": 200.00},
                {"name": "4TH STREET RED GLASS", "price": 250.00},
                {"name": "BLACK LABEL", "price": 200.00},
                {"name": "Gordons", "price": 200.00},
                {"name": "Redlabel", "price": 200.00},
                {"name": "DUREX", "price": 150.00}
            ]
        },
        "SNACKS": {
            "": [
                {"name": "CHOMA SAUSAGE", "price": 100.00},
                {"name": "SAMOSA", "price": 50.00}
            ]
        }
    }
    return inventory

def create_styled_header(ws, headers, row_num=1):
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

def style_data_cells(ws, start_row, end_row, start_col, end_col):
    data_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    data_font = Font(size=10)
    data_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            if row % 2 == 0:
                cell.fill = data_fill
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

def generate_report(sales_data, expenditure):
    config = load_config()
    today = datetime.now()
    filename = os.path.join(DATA_DIR, f"Sales_Report_{today.strftime('%Y%m%d_%H%M%S')}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)
    
    # Sales Summary Sheet
    ws_summary = wb.create_sheet("Sales Summary")
    ws_summary['A1'] = f"{config['bar_name']} - Daily Sales Report"
    ws_summary['A1'].font = Font(size=16, bold=True, color="366092")
    ws_summary['A2'] = f"Date: {today.strftime('%A, %B %d, %Y')}"
    ws_summary['A2'].font = Font(size=12, italic=True)

    headers = ["Product", "Category", "Subcategory", "Stock Before", "Qty Sold", "Price", "Total", "Stock After"]
    create_styled_header(ws_summary, headers, 4)
    
    # Get full inventory
    inventory = get_full_inventory()
    
    # Convert sales data to a dictionary for quick lookup
    sales_dict = {}
    for sale in sales_data:
        product_name = sale["Product Name"].strip().upper()
        sales_dict[product_name] = sale
    
    current_row = 5
    total_qty_sold = 0
    total_sales = 0.0
    
    # Iterate through all inventory items
    for category, subcategories in inventory.items():
        for subcategory, products in subcategories.items():
            for product in products:
                product_name = product["name"].strip()
                if not product_name:  # Skip empty product names
                    continue
                    
                # Check if this product was sold
                product_key = product_name.upper()
                if product_key in sales_dict:
                    sale_data = sales_dict[product_key]
                    stock_before = sale_data.get("Stock Before", 0)
                    qty_sold = sale_data.get("Quantity Sold", 0)
                    price = sale_data.get("Price Per Unit", product["price"])
                    total_sale = sale_data.get("Total Sale", 0)
                    stock_after = sale_data.get("Stock After", stock_before - qty_sold)
                    
                    total_qty_sold += qty_sold
                    total_sales += total_sale
                else:
                    # Product not sold, show with empty sales data
                    stock_before = 0  # You might want to pass actual stock data
                    qty_sold = 0
                    price = product["price"]
                    total_sale = 0
                    stock_after = stock_before
                
                # Write row data
                ws_summary[f'A{current_row}'] = product_name
                ws_summary[f'B{current_row}'] = category
                ws_summary[f'C{current_row}'] = subcategory if subcategory else ""
                ws_summary[f'D{current_row}'] = stock_before
                ws_summary[f'E{current_row}'] = qty_sold
                ws_summary[f'F{current_row}'] = price
                ws_summary[f'G{current_row}'] = total_sale
                ws_summary[f'H{current_row}'] = stock_after
                
                current_row += 1

    # Style all data cells
    if current_row > 5:
        style_data_cells(ws_summary, 5, current_row - 1, 1, 8)

    # Add totals row
    total_row = current_row
    ws_summary[f'A{total_row}'] = "TOTALS"
    ws_summary[f'A{total_row}'].font = Font(bold=True)
    ws_summary[f'E{total_row}'] = total_qty_sold
    ws_summary[f'G{total_row}'] = total_sales
    
    total_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    for col in range(1, 9):
        cell = ws_summary.cell(row=total_row, column=col)
        cell.fill = total_fill
        cell.font = Font(bold=True)

    # Financial Summary Sheet
    ws_financial = wb.create_sheet("Financial Summary")
    total_income = total_sales
    tax_amount = total_income * config.get('tax_rate', 0.16)
    net_income = total_income - expenditure - tax_amount
    
    financial_data = [
        ["Metric", f"Amount ({config['currency']})", "Percentage"],
        ["Gross Revenue", total_income, "100%"],
        ["Less: Expenditure", expenditure, f"{(expenditure/total_income*100):.1f}%" if total_income > 0 else "0%"],
        ["Less: Tax (VAT)", tax_amount, f"{(tax_amount/total_income*100):.1f}%" if total_income > 0 else "0%"],
        ["Net Profit", net_income, f"{(net_income/total_income*100):.1f}%" if total_income > 0 else "0%"]
    ]
    
    ws_financial['A1'] = "Financial Summary"
    ws_financial['A1'].font = Font(size=16, bold=True, color="366092")
    create_styled_header(ws_financial, financial_data[0], 3)
    
    for i, row in enumerate(financial_data[1:], 4):
        ws_financial[f'A{i}'] = row[0]
        ws_financial[f'B{i}'] = row[1]
        ws_financial[f'C{i}'] = row[2]
    
    style_data_cells(ws_financial, 4, 3 + len(financial_data), 1, 3)

    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value is not None)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    wb.save(filename)
    save_analytics_data(sales_data, expenditure, total_income, net_income)
    return filename

def save_analytics_data(sales_data, expenditure, total_income, net_income):
    analytics_file = os.path.join(ANALYTICS_DIR, "daily_analytics.json")
    
    try:
        with open(analytics_file, 'r') as f:
            analytics = json.load(f)
    except FileNotFoundError:
        analytics = []
    
    today_data = {
        "date": datetime.now().strftime("%Y-%m-%d"),
        "total_income": total_income,
        "expenditure": expenditure,
        "net_income": net_income,
        "total_items_sold": sum(row["Quantity Sold"] for row in sales_data),
        "products_sold": len([row for row in sales_data if row["Quantity Sold"] > 0])
    }
    
    # Remove any existing entry for today and add the new one
    analytics = [entry for entry in analytics if entry["date"] != today_data["date"]]
    analytics.append(today_data)
    analytics = sorted(analytics, key=lambda x: x["date"])[-30:]  # Keep last 30 days
    
    with open(analytics_file, 'w') as f:
        json.dump(analytics, f, indent=2)

# Example usage function
def example_usage():
    """Example of how to use the report generator"""
    
    # Example sales data (only products that were actually sold)
    sample_sales_data = [
        {
            "Product Name": "Tusker Cans",
            "Category": "BEERS",
            "Subcategory": "CANS", 
            "Stock Before": 35,
            "Quantity Sold": 2,
            "Price Per Unit": 270,
            "Total Sale": 540,
            "Stock After": 33
        },
        {
            "Product Name": "Heineken Bottle",
            "Category": "BEERS", 
            "Subcategory": "BOTTLES",
            "Stock Before": 20,
            "Quantity Sold": 1,
            "Price Per Unit": 300,
            "Total Sale": 300,
            "Stock After": 19
        },
        {
            "Product Name": "Guinness",
            "Category": "BEERS",
            "Subcategory": "BOTTLES",
            "Stock Before": 25,
            "Quantity Sold": 4,
            "Price Per Unit": 250,
            "Total Sale": 1000,
            "Stock After": 21
        },
        {
            "Product Name": "WhiteCap",
            "Category": "BEERS",
            "Subcategory": "BOTTLES", 
            "Stock Before": 30,
            "Quantity Sold": 2,
            "Price Per Unit": 250,
            "Total Sale": 500,
            "Stock After": 28
        },
        {
            "Product Name": "Tusker Lager",
            "Category": "BEERS",
            "Subcategory": "BOTTLES",
            "Stock Before": 40,
            "Quantity Sold": 3,
            "Price Per Unit": 250,
            "Total Sale": 750,
            "Stock After": 37
        }
    ]
    
    # Example expenditure
    expenditure = 500.0
    
    # Generate the report
    report_file = generate_report(sample_sales_data, expenditure)
    print(f"Report generated: {report_file}")

if __name__ == "__main__":
    example_usage()